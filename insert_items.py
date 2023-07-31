from encrypted import Connection
import openpyxl as xl

class ImportData(Connection):
    def __init__(self):
        print("Loading..")
        self.connector()
    
    def importing(self):
        wb = xl.load_workbook('omar.xlsx')
        sheet1 = wb['Sheet1']
        for row in range(2,sheet1.max_row+1):
            id = sheet1.cell(row,1).value
            name = sheet1.cell(row,2).value
            count =  sheet1.cell(row,3).value
            item_price = sheet1.cell(row,4).value
            price = sheet1.cell(row,5).value
            itemSellPrice = sheet1.cell(row,6).value
            countPerOne = sheet1.cell(row,7).value
            code = sheet1.cell(row,8).value
            expire = sheet1.cell(row,9).value
            big_item = sheet1.cell(row,10).value
            smallItem = sheet1.cell(row,11).value
            itemCost = sheet1.cell(row,12).value
            totalCount = sheet1.cell(row,13).value
            stock = sheet1.cell(row,14).value
            alert = sheet1.cell(row,15).value
            self.cr.execute(F'''INSERT INTO item(id,name , count, item_price, price,itemSellPrice,countPerOne,
                code, expire,big_item,smallItem, itemCost, totalCount,stock,alert) VALUES("{id}", "{name}", "{count}","{item_price}",
                "{price}","{itemSellPrice}","{countPerOne}","{code}","{expire}","{big_item}","{smallItem}","{itemCost}","{totalCount}",
                    "{stock}","{alert}")''')
            self.db.commit()

a = ImportData()
a.importing()